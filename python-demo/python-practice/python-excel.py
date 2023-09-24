from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = [
    {
        'name':'小白',
        'tall':180,
        'age': 23,
        'weight':74
    },
    {
        'name':'小黃',
        'tall':177,
        'age': 26,
        'weight':65
    },
    {
        'name':'小綠',
        'tall':165,
        'age': 18,
        'weight':50
    }
]
wb=Workbook()
ws= wb.active

title=['姓名','身高','年紀','體重']
ws.append(title)

for person in data:
    ws.append(list(person.value()))
    
for col in range(2,5):
    char =get_column_letter(col)
    ws[char +'7']=f'=AVERAGE({char +"2"}:{char + "6"})'
    
for col in range(1,5):
    char =get_column_letter(col)
    ws[char +'1'].font = Font(bold=True)
    
wb.save('data,xlsx')